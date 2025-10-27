from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


def auto_fit_columns(ws):
    """Auto-adjust column widths based on content length."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2


def create_excel_report(
    project,
    vm_data,
    sql_data,
    gke_data,
    owner_data,
    bucket_data,
    fw_data,
    lb_data=None,
    cf_data=None  # ✅ Cloud Functions & Cloud Run data
):
    """Generate a detailed security audit Excel report."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ----------------- Summary Sheet -----------------
    summary_headers = [
        "Project ID",
        "Audit Time",
        "VMs",
        "SQL",
        "GKE",
        "Owners",
        "Buckets",
        "Firewall Rules",
        "Load Balancers",
        "Cloud Functions / Cloud Run"
    ]
    summary.append(summary_headers)
    summary.append([
        project,
        now,
        len(vm_data),
        len(sql_data),
        len(gke_data),
        len(owner_data),
        len(bucket_data),
        len(fw_data),
        len(lb_data) if lb_data else 0,
        len(cf_data) if cf_data else 0
    ])

    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    auto_fit_columns(summary)

    # ----------------- Helper to Add Data Sheets -----------------
    def add_sheet(name, headers, data):
        """Add a formatted sheet with safe data conversion."""
        ws = wb.create_sheet(name)
        ws.append(headers)

        # Header formatting
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4F81BD")
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        if data:
            for row in data:
                if isinstance(row, dict):
                    safe_row = [row.get(h.lower().replace(" ", "_"), "") for h in headers]
                else:
                    safe_row = []
                    for v in row:
                        if isinstance(v, list):
                            safe_row.append(", ".join(map(str, v)))
                        elif isinstance(v, dict):
                            safe_row.append(str(v))
                        else:
                            safe_row.append(v)
                ws.append(safe_row)
        else:
            ws.append(["✅ No issues found."])

        ws.freeze_panes = "A2"
        auto_fit_columns(ws)

    # ----------------- Individual Sheets -----------------
    add_sheet("VMs", ["Instance", "Zone", "Public IP"], vm_data)
    add_sheet("SQL", ["Instance", "Public IP"], sql_data)
    add_sheet("GKE", ["Cluster", "Endpoint", "Private Nodes"], gke_data)
    add_sheet("Owners", ["Member", "Role"], owner_data)
    add_sheet("Buckets", ["Bucket", "Role", "Member"], bucket_data)

    # ----------------- Firewall Rules -----------------
    fw_headers = [
        "Rule Name",
        "Direction",
        "Protocols",
        "Source Ranges",
        "Network",
        "Priority",
        "Disabled"
    ]
    add_sheet("FirewallRules", fw_headers, fw_data)

    # ----------------- Load Balancers -----------------
    if lb_data:
        lb_headers = [
            "Name",
            "Scheme",
            "IP",
            "Target",
            "SSL Policy",
            "SSL Cert Status",
            "HTTPS Redirect",
            "Cloud Armor Policy",
            "Armor Rule Strength",
            "Internal Exposure"
        ]
        add_sheet("LoadBalancers", lb_headers, lb_data)

    # ----------------- Cloud Functions & Cloud Run -----------------
    if cf_data:
        cf_headers = [
            "Service Type",
            "Name",
            "Region",
            "Ingress",
            "Authentication",
            "Service Account",
            "VPC Connector",
            "Allow Internal Only",
            "Public Exposure"
        ]
        add_sheet("CloudFunctions_Run", cf_headers, cf_data)

    # ----------------- Save Workbook -----------------
    path = f"/tmp/{project}_security_audit.xlsx"
    wb.save(path)
    return path
